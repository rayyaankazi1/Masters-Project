"""
signal/validation/sample_for_labeling.py
─────────────────────────────────────────
Draws a stratified random sample of fiscal paragraphs for human validation.

Sampling design:
  • 72 paragraphs total (gives buffer for any you skip)
  • 24 per LLM score class (-1 / 0 / +1)
  • Within each class: stratified by president (proportional, min 4 per president
    where the class has enough observations)
  • Excludes paragraphs > 2500 chars (too long to read comfortably in one sitting)
  • Oversamples LLM-dictionary disagreements (~25% of sample) — most diagnostically
    useful cases for the paper
  • Output is FULLY BLINDED: no president, no date, no LLM score, no reason shown

Output:
  signal/validation/paragraphs_to_label.xlsx   — fill in the 'your_score' column
  signal/validation/labeling_key.csv           — kept private, used by evaluate_labels.py

Usage:
  cd ~/Desktop/Masters-Project
  python signal/validation/sample_for_labeling.py
"""

import os, random
import numpy as np
import pandas as pd

random.seed(42)
np.random.seed(42)

_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
SCORED   = os.path.join(_ROOT, 'data', 'interim', 'paragraphs_llm_scored.csv')
OUT_XLSX = os.path.join(_ROOT, 'signal', 'validation', 'paragraphs_to_label.xlsx')
OUT_KEY  = os.path.join(_ROOT, 'signal', 'validation', 'labeling_key.csv')

# ── Load ─────────────────────────────────────────────────────────────────────
df = pd.read_csv(SCORED, low_memory=False)

# Derive dictionary direction for identifying disagreements
df['dict_dir'] = df['has_hawkish'].astype(int) - df['has_dovish'].astype(int)
df['disagrees'] = (
    ((df['llm_score'] == 1)  & (df['dict_dir'] == -1)) |
    ((df['llm_score'] == -1) & (df['dict_dir'] ==  1))
)

# Filter very long paragraphs (hard to read in a sitting)
df = df[df['text_para'].str.len() <= 2500].copy()
df = df.reset_index(drop=True)
df['para_id'] = df.index  # stable ID for matching back

print(f"Pool after length filter: {len(df):,} paragraphs")
print(df.groupby(['president', 'llm_score']).size().unstack(fill_value=0))

# ── Sampling ─────────────────────────────────────────────────────────────────
TOTAL_PER_CLASS = 24          # 24 × 3 classes = 72 total
DISAGREE_PER_CLASS = 6        # ~25% of each class from disagreement cases

PRES_ORDER = ['Macri', 'AF', 'Milei']

sampled_ids = []

for score in [1, 0, -1]:
    pool = df[df['llm_score'] == score].copy()

    # Split into disagreement and agreement pools
    dis_pool  = pool[pool['disagrees']].copy()
    agr_pool  = pool[~pool['disagrees']].copy()

    # Draw disagreement cases first (up to DISAGREE_PER_CLASS)
    n_dis = min(DISAGREE_PER_CLASS, len(dis_pool))
    dis_sample = dis_pool.sample(n=n_dis, random_state=42)
    sampled_ids.extend(dis_sample['para_id'].tolist())

    # Draw remaining from agreement pool, stratified by president
    n_remaining = TOTAL_PER_CLASS - n_dis
    agr_pool = agr_pool[~agr_pool['para_id'].isin(sampled_ids)]

    # Proportional allocation by president
    pres_counts = agr_pool['president'].value_counts()
    total_agr   = pres_counts.sum()
    allocations = {}
    for p in PRES_ORDER:
        if p in pres_counts:
            allocations[p] = max(1, round(n_remaining * pres_counts[p] / total_agr))
    # Correct rounding drift
    while sum(allocations.values()) > n_remaining:
        allocations[max(allocations, key=allocations.get)] -= 1
    while sum(allocations.values()) < n_remaining:
        allocations[min(allocations, key=allocations.get)] += 1

    for p, n in allocations.items():
        p_pool = agr_pool[agr_pool['president'] == p]
        n = min(n, len(p_pool))
        if n > 0:
            samp = p_pool.sample(n=n, random_state=42)
            sampled_ids.extend(samp['para_id'].tolist())

# ── Build outputs ─────────────────────────────────────────────────────────────
sample = df[df['para_id'].isin(sampled_ids)].copy()

# Shuffle so president / score order is not obvious
sample = sample.sample(frac=1, random_state=99).reset_index(drop=True)
sample['item_no'] = range(1, len(sample) + 1)

print(f"\nFinal sample: {len(sample)} paragraphs")
print("LLM score distribution in sample:")
print(sample['llm_score'].value_counts().sort_index())
print("\nBy president:")
print(sample['president'].value_counts())

# ── Save key (private — do NOT look at this before labeling) ──────────────────
key = sample[['item_no', 'para_id', 'president', 'year_month',
              'llm_score', 'llm_reason', 'disagrees']].copy()
key.to_csv(OUT_KEY, index=False)
print(f"\nKey saved to: {OUT_KEY}")
print("⚠  Do NOT open labeling_key.csv until you have finished labeling.")

# ── Save blinded labeling file ────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Instructions sheet ────────────────────────────────────────────────────
    ws_inst = wb.active
    ws_inst.title = "Instructions"

    instructions = [
        ("HUMAN VALIDATION — FISCAL STANCE LABELING", True),
        ("", False),
        ("Task", True),
        ("Read each paragraph and assign a fiscal stance score:", False),
        ("  +1  Hawkish  — advocates fiscal discipline, deficit reduction, spending cuts,", False),
        ("               emission control, debt reduction, privatisation, austerity", False),
        ("   0  Neutral  — describes fiscal situation without clear advocacy; factual;", False),
        ("               mixed or ambiguous stance", False),
        ("  -1  Dovish   — advocates increased spending, social programmes, public", False),
        ("               investment, wage protection, state expansion, debt renegotiation", False),
        ("", False),
        ("Rules", True),
        ("1. Score based ONLY on what the paragraph says — ignore who you think said it.", False),
        ("2. If a paragraph argues BOTH directions, lean toward whichever is stronger.", False),
        ("3. If genuinely ambiguous, score 0.", False),
        ("4. Do not change your answer after moving on — first instinct is usually right.", False),
        ("5. Aim for ~2 minutes per paragraph. Total time: ~2.5 hours for 72 items.", False),
        ("", False),
        ("How to fill in", True),
        ("Go to the 'Paragraphs' sheet. For each row, type -1, 0, or 1 in column C.", False),
        ("You can also add optional notes in column D.", False),
        ("Save the file when done.", False),
        ("", False),
        ("Valid scores: -1   0   1   (no other values)", False),
    ]

    for i, (text, bold) in enumerate(instructions, start=1):
        cell = ws_inst.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=12)
        else:
            cell.font = Font(size=11)
    ws_inst.column_dimensions['A'].width = 80

    # ── Paragraphs sheet ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Paragraphs")

    headers = ['#', 'Paragraph text', 'your_score (-1 / 0 / 1)', 'notes (optional)']
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Alternating row fills
    fill_a = PatternFill("solid", fgColor="EBF3FB")
    fill_b = PatternFill("solid", fgColor="FFFFFF")

    for idx, row in sample.iterrows():
        r = row['item_no'] + 1
        fill = fill_a if row['item_no'] % 2 == 0 else fill_b

        # Item number
        c1 = ws.cell(row=r, column=1, value=row['item_no'])
        c1.alignment = Alignment(horizontal='center', vertical='top')
        c1.fill = fill

        # Paragraph text
        c2 = ws.cell(row=r, column=2, value=row['text_para'])
        c2.alignment = Alignment(wrap_text=True, vertical='top')
        c2.fill = fill

        # Score (blank — to be filled)
        c3 = ws.cell(row=r, column=3, value=None)
        c3.alignment = Alignment(horizontal='center', vertical='top')
        c3.fill = PatternFill("solid", fgColor="FFF2CC")  # yellow highlight

        # Notes
        c4 = ws.cell(row=r, column=4, value=None)
        c4.fill = fill

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 85
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 30

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUT_XLSX)
    print(f"Labeling file saved to: {OUT_XLSX}")

except ImportError:
    # Fallback to plain CSV if openpyxl not available
    out_csv = OUT_XLSX.replace('.xlsx', '.csv')
    sample[['item_no', 'text_para']].rename(
        columns={'item_no': '#', 'text_para': 'paragraph_text'}
    ).assign(your_score='', notes='').to_csv(out_csv, index=False)
    print(f"openpyxl not found — saved plain CSV to: {out_csv}")

print("\nDone. Open paragraphs_to_label.xlsx, go to the Paragraphs sheet,")
print("and fill in column C with -1, 0, or 1 for each paragraph.")
